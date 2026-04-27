import os
from typing import List
from ..models import ExtractResult, MAX_FULL_TEXT, MAX_TRUNCATED_TEXT


class ExcelExtractor:
    def supported_extensions(self) -> List[str]:
        return ["xlsx"]

    def extract(self, file_path: str) -> ExtractResult:
        try:
            import openpyxl
        except ImportError:
            return ExtractResult.failed("未安装 openpyxl 库，无法提取 .xlsx 文件")

        try:
            file_size = os.path.getsize(file_path)
        except OSError:
            file_size = 0

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            return ExtractResult.failed(f"无法打开 Excel 文件: {e}")

        parts = []
        total_rows = 0
        total_cols = 0

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_parts = [f"## Sheet: {sheet_name}"]
                row_count = 0
                max_col = 0

                for row in ws.iter_rows(values_only=True):
                    if all(c is None for c in row):
                        continue
                    if row_count >= 200:
                        sheet_parts.append(f"... (仅显示前 200 行，共 ? 行)")
                        break
                    cells = [str(c) if c is not None else "" for c in row]
                    max_col = max(max_col, len(cells))
                    sheet_parts.append("| " + " | ".join(cells) + " |")
                    row_count += 1

                total_rows += row_count
                total_cols = max(total_cols, max_col)
                parts.append("\n".join(sheet_parts))
        finally:
            wb.close()
        text = "\n\n".join(parts)
        file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0

        if len(text) <= MAX_FULL_TEXT:
            return ExtractResult.success(text, method="full", metadata={
                "sheet_count": len(wb.sheetnames),
                "total_rows": total_rows,
                "max_columns": total_cols,
                "char_count": len(text),
                "size_bytes": file_size,
            })

        truncated = text[:MAX_TRUNCATED_TEXT]
        return ExtractResult.success(truncated, method="truncated", metadata={
            "sheet_count": len(wb.sheetnames),
            "total_rows": total_rows,
            "max_columns": total_cols,
            "char_count": len(text),
            "displayed_chars": MAX_TRUNCATED_TEXT,
            "size_bytes": file_size,
        })
